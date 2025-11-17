import json
import openpyxl

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Prefetch
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect
from django.urls import reverse
from django.views.generic import ListView, DetailView
from datetime import datetime, timedelta, date
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from urllib.parse import quote

from accounts.models import Profile, Arrangement
from core.models import Department, Office, Position


@login_required
def dashboard(request):
    return render(request, "core/dashboard.html")


@login_required
def main_menu(request):
    return render(request, "core/main-menu.html")


@login_required
def contacts(request):
    return render(request, "core/contacts.html")


@login_required
def worker_list(request):
    return render(request, "core/empl_list.html")


@login_required
def settings(request):
    return render(request, "core/settings.html")


@login_required
def arrangement(request):
    return render(request, "core/arrangement.html")


class EmployeePhonesListView(ListView):
    model = Department
    template_name = "core/contacts.html"
    context_object_name = "departments_data"

    def get_queryset(self):
        office_id = self.request.GET.get("office")
        profiles_qs = Profile.objects.select_related("position", "office")
        query = self.request.GET.get("q")

        if query:
            profiles_qs = profiles_qs.filter(
                Q(last_name__icontains=query) |
                Q(first_name__icontains=query) |
                Q(patronymic__icontains=query) |
                Q(phone_number_work__icontains=query) |
                Q(phone_number_government__icontains=query) |
                Q(phone_number_mobile__icontains=query)
            )

        if office_id:
            profiles_qs = profiles_qs.filter(office_id=office_id)

        positions_qs = Position.objects.prefetch_related(
            Prefetch("profiles", queryset=profiles_qs, to_attr="prefetched_profiles")
        )

        departments_qs = Department.objects.prefetch_related(
            Prefetch("positions", queryset=positions_qs, to_attr="prefetched_positions")
        ).order_by("name")

        return departments_qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        office_id = self.request.GET.get("office")

        departments_list = []

        for dept in context["departments_data"]:
            employees = []
            for pos in getattr(dept, "prefetched_positions", []):
                for prof in getattr(pos, "prefetched_profiles", []):
                    employees.append({"profile": prof, "position": pos})

            if employees:
                departments_list.append({"dept": dept, "employees": employees})

        context["departments_list"] = departments_list
        context["offices"] = Office.objects.all()
        context["selected_office"] = office_id or ""
        context["query"] = self.request.GET.get("q", "")
        return context


class ProfileOfficesListView(ListView):
    model = Profile
    template_name = "core/empl_list.html"
    context_object_name = "employees"

    def get_queryset(self):
        queryset = super().get_queryset().select_related("office")

        # —Ñ–∏–ª—å—Ç—Ä–∞—Ü–∏—è –ø–æ —Ñ–∏–ª–∏–∞–ª—É
        office_id = self.request.GET.get("office")
        if office_id:
            queryset = queryset.filter(office_id=office_id)
        query = self.request.GET.get("q")
        if query:
            queryset = queryset.filter(
                Q(last_name__icontains=query) |
                Q(first_name__icontains=query) |
                Q(email__icontains=query)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["offices"] = Office.objects.all()
        context["selected_office"] = self.request.GET.get("office")
        return context


class EmployeeDetailView(DetailView):
    model = Profile
    template_name = "accounts/dashboard_empl.html"  # –ø—É—Ç—å –∫ —à–∞–±–ª–æ–Ω—É
    context_object_name = "employee"


class ArrangementListView(ListView):
    model = Arrangement
    template_name = "core/arrangement.html"
    context_object_name = "arrangements"
    OFFICE_NAME = "–ë–æ—Ä–±–æ—Ä–¥—É–∫ –∞–ø–ø–∞—Ä–∞—Ç"

    def get_selected_date(self):
        date_str = self.request.GET.get("date")
        if date_str:
            try:
                return timezone.datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                pass
        return timezone.localdate()

    def get_queryset(self):
        selected_date = self.get_selected_date()
        return (
            Arrangement.objects.filter(
                date_create=selected_date,
                profile__office__name=self.OFFICE_NAME
            )
            .select_related("profile", "position")
            .order_by("profile__last_name", "profile__first_name")
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        selected_date = self.get_selected_date()
        qs = self.get_queryset()

        context["selected_date"] = selected_date
        context["previous_date"] = selected_date - timedelta(days=1)
        context["next_date"] = selected_date + timedelta(days=1)

        context["apparatus"] = qs.filter(profile__is_inspector=False)
        context["inspectors"] = qs.filter(profile__is_inspector=True)
        context["is_empty"] = not qs.exists()  # üëà –ü—Ä–æ–≤–µ—Ä—è–µ–º, –µ—Å—Ç—å –ª–∏ –∑–∞–ø–∏—Å–∏

        return context


def generate_arrangement_day(request):
    """–°–æ–∑–¥–∞–Ω–∏–µ –∑–∞–ø–∏—Å–µ–π –¥–ª—è –∫–æ–Ω–∫—Ä–µ—Ç–Ω–æ–≥–æ –¥–Ω—è"""
    if request.method == "POST":
        date_str = request.POST.get("date")
        try:
            date_value = timezone.datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            messages.error(request, "–ù–µ–≤–µ—Ä–Ω–∞—è –¥–∞—Ç–∞.")
            return redirect("arrangement")

        # –û—Ç–±–∏—Ä–∞–µ–º —Ç–æ–ª—å–∫–æ —Ü–µ–Ω—Ç—Ä–∞–ª—å–Ω—ã–π –∞–ø–ø–∞—Ä–∞—Ç
        employees = Profile.objects.filter(office__name="–ë–æ—Ä–±–æ—Ä–¥—É–∫ –∞–ø–ø–∞—Ä–∞—Ç")
        existing_ids = Arrangement.objects.filter(date_create=date_value).values_list("profile_id", flat=True)
        new_records = [
            Arrangement(profile=emp, position=emp.position, date_create=date_value)
            for emp in employees if emp.id not in existing_ids
        ]
        Arrangement.objects.bulk_create(new_records)
        messages.success(request, f"–°—Ñ–æ—Ä–º–∏—Ä–æ–≤–∞–Ω–æ {len(new_records)} –∑–∞–ø–∏—Å–µ–π –Ω–∞ {date_value.strftime('%d.%m.%Y')}.")

        return redirect(f"{reverse('arrangement')}?date={date_value}")


def arrangement_update(request, pk):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            field = data.get("field")
            value = data.get("value")

            arrangements = Arrangement.objects.get(pk=pk)

            # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –µ—Å—Ç—å –ª–∏ —Ç–∞–∫–æ–µ –ø–æ–ª–µ —É –º–æ–¥–µ–ª–∏
            if hasattr(arrangements, field):
                setattr(arrangements, field, value)
                arrangements.save(update_fields=[field])
                return JsonResponse({"success": True, "field": field, "value": value})
            else:
                return JsonResponse({"success": False, "error": "Invalid field"})
        except Arrangement.DoesNotExist:
            return JsonResponse({"success": False, "error": "Object not found"})
    return JsonResponse({"success": False, "error": "Invalid request"})


def import_arrangement_day(request):
    """–ò–º–ø–æ—Ä—Ç –¥–∞–Ω–Ω—ã—Ö –∏–∑ –≤—ã–±—Ä–∞–Ω–Ω–æ–π –¥–∞—Ç—ã (–∫–æ–ø–∏—Ä—É–µ—Ç –≤—Å–µ –ø–æ–ª—è –≤ —Ç–µ–∫—É—â–∏–π –¥–µ–Ω—å)"""
    if request.method == "POST":
        source_date = timezone.datetime.fromisoformat(request.POST.get("source_date")).date()
        target_date = timezone.datetime.fromisoformat(request.POST.get("target_date")).date()

        source_records = Arrangement.objects.filter(date_create=source_date)
        if not source_records.exists():
            messages.warning(request, f"–ù–µ—Ç –¥–∞–Ω–Ω—ã—Ö –∑–∞ {source_date.strftime('%d.%m.%Y')} –¥–ª—è –∏–º–ø–æ—Ä—Ç–∞.")
            return redirect(f"{reverse('arrangement')}?date={target_date}")

        Arrangement.objects.filter(date_create=target_date).delete()

        new_records = [
            Arrangement(
                profile=rec.profile,
                position=rec.position,
                audit_conducting=rec.audit_conducting,
                audit_purpose=rec.audit_purpose,
                order_num_date=rec.order_num_date,
                order_dates=rec.order_dates,
                audit_address=rec.audit_address,
                on_status=rec.on_status,
                time_check=rec.time_check,
                time_not_start=rec.time_not_start,
                date_create=target_date,
            )
            for rec in source_records
        ]
        Arrangement.objects.bulk_create(new_records)
        messages.success(request,
                         f"–ò–º–ø–æ—Ä—Ç–∏—Ä–æ–≤–∞–Ω–æ {len(new_records)} –∑–∞–ø–∏—Å–µ–π –∏–∑ {source_date.strftime('%d.%m.%Y')}.")
        return redirect(f"{reverse('arrangement')}?date={target_date}")


def clear_arrangement_day(request):
    """–û—á–∏—â–∞–µ—Ç –¥–∞–Ω–Ω—ã–µ –≤ —Ç–∞–±–ª–∏—Ü–µ –∑–∞ –¥–µ–Ω—å"""
    if request.method == "POST":
        date_value = timezone.datetime.fromisoformat(request.POST.get("date")).date()
        Arrangement.objects.filter(date_create=date_value).update(
            audit_conducting="",
            audit_purpose="",
            order_num_date="",
            order_dates="",
            audit_address="",
            on_status="",
            time_check="",
            time_not_start="",
        )
        messages.info(request, f"–î–∞–Ω–Ω—ã–µ –∑–∞ {date_value.strftime('%d.%m.%Y')} –æ—á–∏—â–µ–Ω—ã.")
    return redirect(f"{reverse('arrangement')}?date={date_value}")


def export_contacts_excel(request):
    """–≠–∫—Å–ø–æ—Ä—Ç —Å–ø–∏—Å–∫–∞ —Å–æ—Ç—Ä—É–¥–Ω–∏–∫–æ–≤ –≤ Excel"""
    # –°–æ–∑–¥–∞–µ–º –Ω–æ–≤—É—é –∫–Ω–∏–≥—É
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    offices = (
        Profile.objects.filter(office__isnull=False)
        .select_related("office", "position", "position__department")
        .order_by("office__name", "position__department__name", "last_name")
    )

    # –ì—Ä—É–ø–ø–∏—Ä—É–µ–º —Å–æ—Ç—Ä—É–¥–Ω–∏–∫–æ–≤ –ø–æ –æ—Ñ–∏—Å–∞–º
    offices_data = {}
    for p in offices:
        offices_data.setdefault(p.office.name, []).append(p)

    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )

    today_str = timezone.now().strftime("%d.%m.%Y")

    # –ó–∞–≥–æ–ª–æ–≤–∫–∏
    for office_name, profiles in offices_data.items():
        ws = wb.create_sheet(title=office_name[:31])  # Excel –æ–≥—Ä–∞–Ω–∏—á–∏–≤–∞–µ—Ç –¥–ª–∏–Ω—É –∏–º–µ–Ω–∏ –ª–∏—Å—Ç–∞

        # –ó–∞–≥–æ–ª–æ–≤–æ–∫
        ws.merge_cells("A1:G1")
        ws["A1"] = (
            "–ö—ã—Ä–≥—ã–∑ –†–µ—Å–ø—É–±–ª–∏–∫–∞—Å—ã–Ω—ã–Ω –≠—Å–µ–ø—Ç”©”© –ø–∞–ª–∞—Ç–∞—Å—ã–Ω—ã–Ω "
            f"{office_name} –∫—ã–∑–º–∞—Ç–∫–µ—Ä–ª–µ—Ä–∏–Ω–∏–Ω —Ç–µ–ª–µ—Ñ–æ–Ω–¥–æ—Ä—É–Ω—É–Ω –∂–∞–Ω–∞ –æ—Ç—É—Ä–≥–∞–Ω –∫–∞–±–∏–Ω–µ—Ç—Ç–µ—Ä–∏–Ω–∏–Ω –º–∞–∞–ª—ã–º–¥–∞–º–∞—Å—ã"
        )
        ws["A1"].font = Font(name="Times New Roman", size=14, bold=True, color="AA0000")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 40

        # –ü–æ–¥–∑–∞–≥–æ–ª–æ–≤–æ–∫ (–¥–∞—Ç–∞)
        ws.merge_cells("A2:G2")
        ws["A2"] = f"({timezone.now().strftime('%d.%m.%Y')}-–∂. –∫–∞—Ä–∞—Ç–∞)"
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A2"].font = Font(name="Times New Roman", size=12, italic=True, color="555555")

        if "–ë–æ—Ä–±–æ—Ä–¥—É–∫ –∞–ø–ø–∞—Ä–∞—Ç" in office_name:
            ws.merge_cells("A3:G3")
            ws["A3"] = (
                "–ø–æ—á—Ç–∞–Ω—ã–Ω –¥–∞—Ä–µ–≥–∏: 720033, –ë–∏—à–∫–µ–∫ —à., –ò—Å–∞–Ω–æ–≤ –∫”©—á., 131, —Ñ–∞–∫—Å: 32 35 11"
            )
            ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A3"].font = Font(name="Times New Roman", size=12, italic=True, color="AA0000")
            start_row = 5
        else:
            start_row = 4

        # –ó–∞–≥–æ–ª–æ–≤–∫–∏ —Ç–∞–±–ª–∏—Ü—ã
        headers = ["‚Ññ", "–ê—Ç—ã-–∂”©–Ω“Ø", "–ö—ã–∑–º–∞—Ç –æ—Ä–¥—É", "–ö—ã–∑–º–∞—Ç—Ç—ã–∫ —Ç–µ–ª–µ—Ñ–æ–Ω ‚Ññ", "”®–∫–º”©—Ç—Ç“Ø–∫ ‚Ññ", "–ú–æ–±–∏–ª–¥–∏–∫ —Ç–µ–ª–µ—Ñ–æ–Ω ‚Ññ", "–ö–∞–±. ‚Ññ"]
        ws.append(headers)
        header_row = start_row

        header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=header)
            cell.font = Font(name="Times New Roman", size=12, bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        # –ù–∞—á–∏–Ω–∞–µ–º —Å–æ —Å—Ç—Ä–æ–∫–∏ 5
        row_num = header_row + 1

        departments = {}
        for p in profiles:
            dept_name = p.position.department.name if p.position and p.position.department else "–ë–∞—à–∫–∞"
            departments.setdefault(dept_name, []).append(p)

        # –î–æ–±–∞–≤–ª—è–µ–º –ø–æ–¥—Ä–∞–∑–¥–µ–ª—ã
        for dept, profs in departments.items():
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
            cell = ws.cell(row=row_num, column=1, value=dept)
            cell.font = Font(name="Times New Roman", size=12, bold=True, color="000080")
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            row_num += 1

            for i, p in enumerate(profs, start=1):
                ws.append([
                    i,
                    p.full_name(),
                    p.position.title if p.position else "",
                    p.phone_number_work or "",
                    p.phone_number_government or "",
                    p.phone_number_mobile or "",
                    p.office_number or "",
                ])
                for col in range(1, 8):
                    c = ws.cell(row=row_num, column=col)
                    c.font = Font(name="Times New Roman", size=12)
                    c.border = border
                    c.alignment = Alignment(vertical="center", wrap_text=True)
                row_num += 1

            ws.append([])
            row_num += 1

        # –ü–æ–¥–≥–æ–Ω —à–∏—Ä–∏–Ω—ã
        widths = [5, 30, 20, 18, 15, 20, 10]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    filename = f"–°–ø—Ä–∞–≤–æ—á–Ω–∏–∫ —Ç–µ–ª–µ—Ñ–æ–Ω–æ–≤ –Ω–∞ {today_str}.xlsx"
    encoded_filename = quote(filename)

    # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —Ñ–∞–π–ª –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{encoded_filename}"
    wb.save(response)
    return response
